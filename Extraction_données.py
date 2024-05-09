import cv2
import pytesseract
import fitz  # PyMuPDF
import pdfplumber
import re
from collections import namedtuple
import pandas as pd
import streamlit as st 
from request import *
import os 
import pandas as pd
from PIL import Image
import json
import openpyxl


chemin_tesseract=r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.tesseract_cmd=chemin_tesseract

import uuid
st.title("Extraction des Données")
logo_path = "gproconsulting.png" 
st.sidebar.image(logo_path, use_column_width=True)
type_facture = st.sidebar.selectbox("Sélectionner le type de facture", ["DESIPRO PTE LTD", "Société Sublitunis", "DECOTHLON PRODU ZIONE ITALIA", "VTL", "CARVICO","JERSEY LOMELLINA","Orine chine","origine Turque"],index=None,
   placeholder="Select facture type..",)

def choisir_fichier():
    fichier = st.file_uploader("Choisissez un fichier", type=["jpg", "jpeg", "png", "pdf", "xlsx", "xls"])
    chemin_fichier = None  # Initialisation de la variable
    if fichier is not None:
        os.makedirs('temp', exist_ok=True)
        chemin_fichier = os.path.join('temp', fichier.name)
        with open(chemin_fichier, 'wb') as f:
            f.write(fichier.getbuffer())
        
        # Afficher le fichier téléchargé
        if fichier.type.startswith('image'):
            st.image(fichier, caption='Fichier choisi')
        elif fichier.type == 'application/pdf':
            st.markdown(f"Téléchargement du fichier [{fichier.name}]({chemin_fichier})")
        elif fichier.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or fichier.type == 'application/vnd.ms-excel':
            st.markdown(f"Téléchargement du fichier [{fichier.name}]({chemin_fichier})")
        else:
            st.write(f"Type de fichier non pris en charge: {fichier.type}")
        
    return chemin_fichier


def extraire_informations_document(chemin_fichier):
    informations_extraites = None  
    if chemin_fichier is not None:
        if chemin_fichier.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
            # Traitement pour les fichiers image
            informations_extraites = extraire_informations_image(chemin_fichier)     
        elif chemin_fichier.lower().endswith('.pdf'):
            informations_extraites = extraire_informations_pdf(chemin_fichier)
        elif chemin_fichier.lower().endswith(('.xlsx', '.xls')):
            informations_extraites = extraire_informations_excel(chemin_fichier)
        else:
            st.error("Format de fichier non pris en charge")
    else:
        st.warning("Aucun fichier sélectionné")
    return informations_extraites



def extraire_informations_image(chemin_image):
    # Lecture de l'image
    image = cv2.imread(chemin_image)
    # Conversion en niveaux de gris
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Application d'un filtre gaussien pour réduire le bruit
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    # Binarisation de l'image avec seuillage d'Otsu
    _, thresholded = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # Recherche de contours
    contours, _ = cv2.findContours(thresholded, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # Initialisation de la variable pour stocker le texte extrait
    texte_brut = ""
    
    # Parcours des contours
    for contour in contours:
        # Filtre des contours basés sur leur taille
        if cv2.contourArea(contour) > 100:
            # Extraction des coordonnées du rectangle englobant
            x, y, w, h = cv2.boundingRect(contour)
            
            # Extraction de la région d'intérêt (ROI)
            roi = image[y:y+h, x:x+w]
            texte_extrait = pytesseract.image_to_string(roi, lang='eng')
            texte_brut += texte_extrait 
    informations = {"texte": texte_brut}
    collection1.insert_one(informations)
    return informations



def extraire_informations_pdf(chemin_pdf):
    texte_extrait = None  
    try:
        with pdfplumber.open(chemin_pdf) as pdf:
            page = pdf.pages[0]
            texte_extrait = page.extract_text()
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Le fichier PDF '{chemin_pdf}' n'a pas été trouvé.") from e
    except Exception as e:
        raise Exception(f"Une erreur est survenue lors de l'extraction des informations du PDF '{chemin_pdf}'.") from e
    informations_extraites = {
        "texte": texte_extrait,
    }
    collection1.insert_one(informations_extraites)
    return informations_extraites


def extraire_informations_excel(chemin_excel):
    try:
        if type_facture == "origine Turque":
            wb = openpyxl.load_workbook(chemin_excel)

            # Accéder à la première feuille (ou unique) du classeur
            feuille = wb.active

            # Variable pour suivre si l'erreur "too many values to unpack" s'est produite
            erreur_unpack = False
            last_row = feuille.max_row

            # Déterminer l'avant-dernière ligne
            avant_derniere_ligne = last_row - 1
            # Parcourir chaque ligne à partir de la ligne 8
            for row in feuille.iter_rows(min_row=8, max_row=avant_derniere_ligne, values_only=True):
        
                try:
                    # Attribuer chaque valeur à une variable différente
                    col1,col2, col3, col4, col5, col6, col7, col8, col9, col10, col11, col12 = row
                    # Afficher les valeurs stockées dans chaque variable
                    Nom = "origine turque"
                    st.write(str(col6), str(col9), str(col11))
                    data={
                        "Nom":Nom,
                        "type_p":str(col6),
                        "metrage":str(col9),
                        "kilometrage":str(col11)
                    }
                    collection.insert_one(data)
                except ValueError as e:
                    if "too many values to unpack" in str(e) and not erreur_unpack:
                        st.error("Le nombre de colonnes dans le fichier Excel ne correspond pas au format attendu.")
                        erreur_unpack = True
                    elif "too many values to unpack" in str(e) and erreur_unpack:
                        pass  # Ignorer les autres occurrences de l'erreur "too many values to unpack"
                    else:
                        raise e  # Laisser les autres erreurs ValueError être traitées normalement
        else:
            st.warning("Désolé, veuillez vérifier le type de facture")
            return None
    except FileNotFoundError:
        st.error("Le fichier Excel spécifié est introuvable.")
        return None
    except Exception as e:
        st.error(f"Une erreur s'est produite lors de l'extraction des informations du fichier Excel : {str(e)}")
        return None
    finally:
        try:
            # Fermer le fichier Excel
            wb.close()
        except UnboundLocalError:
            pass  # Si wb n'est pas défini, cela signifie qu'il n'a jamais été ouvert



def traitement_Sublitunis_pdf(informations):
    informations_extraites =informations["texte"]
    nom_pattern =  r"Sociéte\s+(\w+)"
    pieces_pattern=r"\b\d+\s*R[A-Za-z]?\d*\b"
    metrage_pattern=r"\b\d+\s(?=\d{2}/\d{2}/\d{4})"
    date_pattern = r"Date : (\d{2}-\d{2}-\d{4})"
    date_pro_pattern = r"\b\d+\s(\d{2}/\d{2}/\d{4})\b"
    num_commande_pattern = r"N°commande : (\d+)"
    destinataire_pattern = r"Destinataire : (\w+)"
    # Extraction des informations avec les expressions régulières
    nom_match = re.search(nom_pattern, informations_extraites)
    date_match = re.search(date_pattern, informations_extraites)
    num_commande_match = re.search(num_commande_pattern, informations_extraites)
    destinataire_match = re.search(destinataire_pattern, informations_extraites)
    pieces_match= re.findall(pieces_pattern, informations_extraites)
    date_pro_match=re.findall(date_pro_pattern , informations_extraites)
    metrage_match=re.findall(metrage_pattern,informations_extraites)
    expression = r"Imprimé\s*:\s*([^\d]+)"

    resultat = re.search(expression, informations_extraites)
    destinataire = None  # Initialisation de la variable destinataire

    if destinataire_match:
        destinataire = destinataire_match.group(1)
    
    if nom_match and date_match and num_commande_match and destinataire_match:
        nom = nom_match.group(1)
        date = date_match.group(1)
        num_commande = num_commande_match.group(1)
        destinataire = destinataire_match.group(1)
        n_pieces= pieces_match
        date_pro=date_pro_match
        metrage=metrage_match
       
        st.write( nom)
        if destinataire=="Enfavet":
            for n_piece,date_pr,metrag in zip(n_pieces,date_pro,metrage):
                data = {
                "Nom": nom,
                "Numéro de commande": num_commande,
                "Date": date_pr,
                "Destinataire":destinataire,
                "n_pieces":n_piece,
                "date_pro":date,
                "metrage":metrag,
                "type_p":resultat.group(1).strip()
            }
                    
                collection.insert_one(data)
    else:
     st.write("verifier le type de facture")         
    

def traitement(chemin_fichier, type_facture, informations_extraites):
    if chemin_fichier.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
        if type_facture == "Société Sublitunis":
            traitement_Sublitunis_image(informations_extraites)
        elif type_facture == "CARVICO":
            traitement_carvico_df(informations_extraites)
        elif type_facture == "VTL":
            traitement_vtl_image(informations_extraites)
        elif type_facture == "DESIPRO PTE LTD":
             traitement_desipro_chine_pdf(informations_extraites)
        elif type_facture == "DECOTHLON PRODU ZIONE ITALIA":
            traitement_DECOTHLON_PRODU_ZIONE_ITALIA_pdf(informations_extraites)
        elif type_facture=="JERSEY LOMELLINA":
            traitement_JERSEY_LOMELLINA_image(informations_extraites)
        elif type_facture=="Orine chine":
            traitement_originechine(informations_extraites)
        else:
            st.error("Type de facture non pris en charge pour les images")
        
    elif chemin_fichier.lower().endswith('.pdf'):
        if type_facture == "Société Sublitunis":
            st.warning("Le type de facture sélectionné est 'Société Sublitunis', veuillez vérifier si c'est correct.")
            traitement_Sublitunis_pdf(informations_extraites)
        elif type_facture == "CARVICO":
            traitement_carvico_df(informations_extraites)
        elif type_facture == "VTL":
           traitement_vtl_pdf(informations_extraites)
        elif type_facture == "DESIPRO PTE LTD":
           traitement_desipro_chine_pdf(informations_extraites)  
        elif type_facture == "DECOTHLON PRODU ZIONE ITALIA":
            traitement_DECOTHLON_PRODU_ZIONE_ITALIA_pdf(informations_extraites)
        elif type_facture=="JERSEY LOMELLINA":
            traitement_JERSEY_LOMELLINA_image(informations_extraites)
        elif type_facture=="Orine chine":
            traitement_originechine(informations_extraites)
        else:
            st.error("Type de facture non pris en charge pour les PDF")
    else:
        st.error("Format de fichier non pris en charge")



def traitement_Sublitunis_image(informations):
        informations_extraites =informations["texte"]
        nom_pattern =  r"Societe\s+(\w+)"
        pieces_pattern=r"\b\d+\s*R[A-Za-z]?\d*\b"
        metrage_pattern=r"(\d+)\s+\d{2}/\d{2}/\d{4}"
        date_pattern = r"Date :\s*(\d{2}-\d{2}-\d{4})"
        date_pro_pattern = r"\b\d{2}/\d{2}/\d{4}\b"
        num_commande_pattern = r"N°commande : (\d+)"
        destinataire_pattern = r"Destinataire : (\w+)"
    # Extraction des informations avec les expressions régulières
        nom_match = re.search(nom_pattern, informations_extraites)
        date_match = re.search(date_pattern, informations_extraites)
        num_commande_match = re.search(num_commande_pattern, informations_extraites)
        destinataire_match = re.search(destinataire_pattern, informations_extraites)
        pieces_match= re.findall(pieces_pattern, informations_extraites)
        date_pro_match=re.findall(date_pro_pattern , informations_extraites)
        metrage_match=re.findall(metrage_pattern,informations_extraites)
        expression =  r"Imprimé\s*:\s*([^\d]+Ey)\s"

        resultat = re.search(expression, informations_extraites)
        if resultat:
            st.write(resultat.group(1).strip(" Ey"))
        else:
            print("Aucune correspondance trouvée.")
        if nom_match:
            nom = nom_match.group(1)
            st.write(nom)
        else:
            nom = None
            alerte(nom) 
        if date_match:
            date = date_match.group(1)
        else:
            date = None
            alerte(date)
        if num_commande_match:
            num_commande = num_commande_match.group(1)
        else:
            num_commande = None
            alerte(num_commande)
        if destinataire_match:
            destinataire = destinataire_match.group(1)
        else:
                destinataire = None
                alerte(destinataire)
        if pieces_match:
            n_pieces= pieces_match
        else:
            n_pieces=None
            alerte(n_pieces)
        if date_pro_match:
            date_pro=date_pro_match
        else:
            date_pro=None
            alerte(date_pro)
        if metrage_match:
            metrage=metrage_match
        else:
            metrage=None
            alerte(metrage)

        # Affichage des résultats
        if n_pieces!=None and date_pro  and nom  :
            for n_piece,date_pr,metrag in zip(n_pieces,date_pro,metrage):
                data = {
       "Nom": nom,
       "Numéro de commande": num_commande,
       "Date": date_pr,
       "Destinataire":destinataire,
       "n_pieces":n_piece,
       "date_pro":date,
       "metrage":metrag,
       "type_p":resultat.group(1).strip(" Ey")
       
}
                collection.insert_one(data)
        else:
                st.write("verifier la facture et le type de facture ")
def traitement_DECOTHLON_PRODU_ZIONE_ITALIA_pdf(informations):
    informations_extraites =informations["texte"]
    modele = r'TOTALE KG\s*=\s*(\d+)'
    tableau_match = re.findall(r'(\d+)\s+(\d+)', informations_extraites)
    nom_match = informations_extraites.splitlines()

    patter_date = r"DATA PACKINGE LIST = (\d+ [A-Z]+ \d+)"
    matches = re.findall(patter_date, informations_extraites)
    expression_motif = r"\b\d{2} (\w+) \d{2}\b"
    data_packinge_list =None
# Recherche du motif dans le texte
    resultats_motif = re.findall(expression_motif, informations_extraites)
    if resultats_motif:
        print("Chaînes de caractères trouvées:")
        for type in resultats_motif:
            type_p=type

    if matches and type_p:
        data_packinge_list = matches[0]
        st.write(type_p)
        st.write("DATA PACKINGE LIST:", data_packinge_list)
        

    else:
       alerte("Aucune correspondance trouvée pour la DATA PACKINGE LIST.")
    colonnes = list(zip(*tableau_match))
   
    if colonnes :
        Mètres = colonnes[0]
        Kgs= colonnes[1]
    else:
        kg_metre=None
        alerte("totale en "+kg_metre)
    if nom_match :
        nom=nom_match[0]
        st.write(nom)
    
    Mètre = Mètres 
    
    
    
   

    if  nom and Mètre and data_packinge_list!=None:
        for Mètre , Kgs in zip(Mètre , Kgs) :
            data = {
                                "Nom": nom,
                                "metrage":Mètre,
                                "kilometrage":Kgs ,
                                "Date":data_packinge_list,
                                "type_p":type_p
                            }
            collection.insert_one(data)
    else:
          st.write("verifier la facture et le type de facture ")
    
    
def traitement_desipro_chine_pdf(informations):
    informations_extraites =informations["texte"]
    name_match = re.search(r'(DESIPRO PTE LTD)', informations_extraites)
    descriptions = re.findall(r'\b\d{1,2} ?\(\d+\) [A-Za-z]\d+\b', informations_extraites)
    expression = r"COLOR\s(.*?)\("

# Recherche de la correspondance dans le texte
    type_p = re.search(expression, informations_extraites)

 
        
   
    if name_match:
        company_name = name_match.group(1)
        st.write("Company Name:", company_name)
        st.write(type_p.group(1))
        
        quantity = re.findall(r'\b\d{1,2} ?YRD', informations_extraites)
        quantities_sans_YRD = [quantite.replace(" YRD", "") for quantite in quantity]
        facteur_conversion = 0.9144

        # Conversion des quantités en mètres
        quantities_metres = [int(qty) * facteur_conversion for qty in quantities_sans_YRD]
        quantities_metres = [str(m) for m in quantities_metres]
        # Affichage des quantités en mètres
        st.write("Quantités converties en mètres:", quantities_metres)
    
        if company_name!=None:
            for metres ,descriptions in  zip(quantities_metres ,descriptions):
                
            
                data = {
        "Nom":company_name ,
        "description":descriptions,
        "metrage":metres,
        "type_p":type_p.group(1)
        
    }   
        
                collection.insert_one(data)
    else:
        st.write("traitement non effectué verifier la facture")
def traitement_vtl_image(informations):

    informations_extraites =informations["texte"]
    pattern =  r"^(.*?)\n"
    nom = 'SVTL'
    pattern_date = r"\b\d{1,2}\s*/\s*\d{1,2}\s*/\s*\d{4}\b"
   # nom = re.search(pattern, informations_extraites)
    date = re.search(pattern_date, informations_extraites)
    pattern_table = r'(R\d+)\s+(\d+)\s+([\d\s,]+)\s+(\d+)'

# Recherche et extraction des informations
    table = re.findall(pattern_table,informations_extraites)

    if date:
        date_livraison = date.group(0)
        st.write("Date de livraison:", date_livraison)
    else:
        print("Date de livraison non trouvée.")
    if nom:
        nom = nom
        st.write("Nom:", nom)
    else:
        st.warning("Nom de l'entreprise non trouvée.")
    
    # Affichage des résultats
    for match in table:
        st.write("N° Pièces:", match[0])
        st.write("Qté/m:", match[1])
        st.write("Qté/kg:", match[2])
        if  nom and match[0] and match[2]: 
            data = {
       "Nom": nom,
       "n_pieces":match[0],
       "Date":date_livraison,
       "kilometrage":match[2],
       
       
       "metrage":match[1]
}
        if nom and date and match[2] :
            collection.insert_one(data)
        else:
          st.write("verifier la facture et le type de facture ")


def traitement_vtl_pdf(informations):


    informations_extraites =informations["texte"]
    pattern =  r"^(.*?)\n"
    pattern_date = r"\b\d{1,2}\s*/\s*\d{1,2}\s*/\s*\d{4}\b"
    nom = 'SVTL'
    date = re.search(pattern_date, informations_extraites)
    pattern_table = r'(R\d+)\s+(\d+)\s+([\d\s,]+)\s+(\d+)'
    
    

    table = re.findall(pattern_table,informations_extraites)

    if date:
        date_livraison = date.group(0)
        st.write("Date de livraison:", date_livraison)
    else:
        print("Date de livraison non trouvée.")
    if nom:
        nom = nom
        st.write("Nom:", nom)
        
    else:
        st.warning("Nom de l'entreprise non trouvée.")
    
    # Affichage des résultats
    for match in table:
        st.write("N° Pièces:", match[0])
        st.write("Qté/m:", match[1])
        st.write("Qté/kg:", match[2])
        
        if  nom and match[0] and match[2]: 
            data = {
       "Nom": nom,
       "n_pieces":match[0],
       "Date":date_livraison,
       "kilometrage":match[2],
      
       "metrage":match[1],
       "type_p":"BLACK N07A"
}
        if nom and date and match[2] :
            collection.insert_one(data)
        else:
          st.write("verifier la facture et le type de facture ")
def traitement_JERSEY_LOMELLINA_image(informations):
    informations_extraites =informations["texte"]
    tableau_pattern = r'\b\d+\s[\d,]+\s[\d,]+'
    tableau_match = re.findall(tableau_pattern, informations_extraites)
    date_pattern = r'(\d{2}/\d{2}/\d{2})'
    date_match = re.search(date_pattern, informations_extraites)
    date = date_match.group(1) if date_match else None

    nom_pattern = r'JERSEY\sLOMELLINA\s*'
    pattern = r'(\d+)\s+(\d+,\d+)\s+(\d+,\d+)'

    expression = r"PIACENZA\*\*\*(.*?)(?=PA)"

# Recherche de la correspondance dans le texte
    resultat = re.search(expression, informations_extraites, re.DOTALL)

    if resultat:
        ligne_trouvee = resultat.group(1).strip()
        # Expression régulière pour extraire uniquement les mots qui commencent par une lettre majuscule suivie de lettres majuscules ou minuscules et d'espaces
        motif = r"\b[A-Za-z]+\b(?:\s[A-Za-z]+)*"
        resultat_final = re.search(motif, ligne_trouvee)
        if resultat_final:
            st.write("tyepe de produit :", resultat_final.group())
            type_p=resultat_final.group()
    

    donnees = re.findall(pattern, informations_extraites)
    from datetime import datetime


    date_obj = datetime.strptime(date, "%d/%m/%y")

    # Vérifier si l'année est à deux chiffres
    if len(date.split("/")[-1]) == 2:
       if date_obj.year < 50:
        date_obj = date_obj.replace(year=date_obj.year + 2000)
    else:
        date_obj = date_obj.replace(year=date_obj.year + 1900)


    # Affichage de la date sous le même format "25/05/2023"
    date_formattee = date_obj.strftime("%d/%m/%Y")
    st.write("Date formatée :", date_formattee)



    nom_match = re.search(nom_pattern,informations_extraites)
    nom = nom_match.group(0).strip() if nom_match else None

    st.write(nom)
    
    if nom and date and tableau_match:
        for donnee in donnees:
            
            data = {
       "Nom": nom,
       "Date":date_formattee,
     
       "kilometrage":donnee[2],
       "metrage":donnee[1],
       "Roll_no":donnee[0],
       "type_p":type_p
}
            collection.insert_one(data)
    else:
        alerte("Nombre d'éléments invalide :", )
def traitement_originechine(informations):
    informations_extraites =informations["texte"]
    lignes = informations_extraites.splitlines()
    
    deuxieme_ligne = lignes[0].rstrip('.')
    st.write(deuxieme_ligne.strip())

    regex_pattern = r"DATE: (\w+\.\d+\w{2},\d{4})"
    resultat = re.search(regex_pattern, informations_extraites)
    metrage= re.findall(r'\b\d{2},\d{2}\b$', informations_extraites, re.MULTILINE)
    kilograme = re.findall(r'\b\d{1,2},\d\b(?= 1\b)', informations_extraites)
    motif = r'(\d{2}/\d{2}/\d{4}) (\d+) (\d+) (\d+) (\d+) (\d+) (\d+) ([\d,.]+) ([\d,.]+)'
    match = re.search(r'Date.*?\n(.*?)\nTOTAL', informations_extraites, re.DOTALL)
    date_facture_match = re.search(r'DATE: (\w+\.\d+(?:TH)?)\,\d{4}',informations_extraites)
   
    details_articles_matches = re.findall(r'(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+\$\s+(\d+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+(\d+)',informations_extraites)

# Extraire le numéro de facture
    numero_facture_match = re.search(r'INVOICE NO\. (\d+)',informations_extraites)
    expression = r"(?<=all\s)(.*?)(?=Y\d+)"

# Recherche de toutes les correspondances dans le texte
    type_p= re.findall(expression, informations_extraites)
    metrag_en_chaine = [str(m) for m in metrage]
    pattern_date= r"DATE NO:(\d{4}/\d{2}/\d{2})"
    correspondance_date = re.search(pattern_date, informations_extraites)
    pattern_metrage = r"KG.*"
    correspondance_metrage= re.search(pattern_metrage, informations_extraites, re.DOTALL)   
   
    if resultat and metrag_en_chaine:
        date = resultat.group(1)
        st.write(date)
        st.write(metrag_en_chaine)
        st.write(kilograme)
        st.write(type_p)
        for kilograme,metrag ,type in zip(kilograme,metrag_en_chaine,type_p):
            data = {
       "Nom": deuxieme_ligne,
       
       "Date": date,
       "kilometrage":kilograme,
       "metrage":metrag,
       "type_p":type
}
            collection.insert_one(data)
    elif match:
            ligne_entre_date_et_total = match.group(1)
            matches = re.findall(motif, ligne_entre_date_et_total)
            if matches:
                for m in matches:
                    st.write("Date:", m[0])
                    st.write("Invoice No:", m[1])
                    st.write("Item Code:", m[2])
                    st.write("Model Code:", m[3])
                    st.write("Supplier No:", m[4])
                    st.write("DE NO:", m[5])
                    st.write("QTY(Y):", m[6])
                    st.write("Unit price:", m[7])
                    st.write("FOB(SH):", m[8])
                    
                    st.write("\n")
                    data={'Nom':deuxieme_ligne,
                          'Date':m[0],
                          "QTY(Y)":m[6],
                          "Unit price": m[7],
                          "FOB(SH)":m[8],
                          "numero_f":m[1]

                    }
                    collection2.insert_one(data)
    elif  date_facture_match and numero_facture_match:
        date_facture = date_facture_match.group(1) 
        numero_facture = numero_facture_match.group(1) 
        details_articles = []
        for match in details_articles_matches:
            details_articles.append({
                'FG_supplier_PO': match[0],
                'DKT_PO': match[1],
                'Model_Description': match[2],
                'Item_Color': match[3],
                'QTY_yds': match[4],
                'QTY_Mtr': match[5],
                'UNIT_PRICE_Yds': match[6],
                'AMOUNT': match[7],
                'N.W': match[8],
                'G.W': match[9],
                'NO.OF_ROLL': match[10]
            })
            data={
                "Nom":deuxieme_ligne,
                "Date":date_facture,
                "numero_f":numero_facture,
                "QTY(Y)": match[4],
                "QTY_Mtr":match[5],
                "Unit price":match[6]

            }
        collection2.insert_one(data)

        st.write("Date de la facture:", date_facture)
        st.write("Numéro de facture:", numero_facture)
        st.write("Détails des articles:")
        for i, article in enumerate(details_articles, start=1):
            st.write("Article", i, ":", article)
    elif correspondance_metrage and correspondance_date:
        lignes =informations_extraites[correspondance_metrage.start():].split('\n')
        for ligne in lignes:
            mots = ligne.strip().split()
            if len(mots) >= 3: 
                
                variable11=mots[10]
                variable12 = mots[11]
                
                variable15=mots[14]
                variable16= mots[15]
                var=variable11+" "+variable12
                date_no = correspondance_date.group(1)
                st.write("Date DATE NO:", date_no)
                st.write(var , variable15,variable16)
                data = {
                        "Nom": deuxieme_ligne,
                        
                        "Date": date_no,
                        "kilometrage":variable16,
                        "metrage":variable15,
                        "type_p":var
                    }
                collection.insert_one(data)

def traitement_carvico_df(informations):
    informations_extraites =informations["texte"]
    pattern_date = r"\b\d{2}/\d{2}/\d{4}\b"
    date = re.findall(pattern_date, informations_extraites)
    pattern_type = r"\*{3}.*?((?:\b\w+\b\s+){2}\b\w+\b)\s*PA"
    pattern = r"fault\n((?:\d.*\n)+)"
    correspondances = re.findall(pattern, informations_extraites)
    # Recherche de correspondance dans le texte
    type_p= re.search(pattern_type, informations_extraites, re.DOTALL)
    Nom='Carvico'
    # Affichage de la correspondance
    if type_p:
        type_pro = type_p.group(1)
        st.write(type_pro)
    else:
        st.write("Aucune correspondance trouvée.")
    if date:
        st.write("Date trouvée :", date[0])
    else:
        st.write("Aucune date trouvée dans le texte.")
    for correspondance in correspondances:
        lignes = correspondance.strip().split('\n')
        for ligne in lignes:
            mots = ligne.split()
            if len(mots) >= 2 and type_pro:
                premiere_valeur = mots[0]
                deuxieme_valeur = mots[1]
                troisieme_valeur = mots[2]
                quatrieme_valeur = mots[3]
                cinquieme_valeur = mots[4]
                
                st.write("Quatrième valeur :", quatrieme_valeur)
                st.write("Cinquième valeur :", cinquieme_valeur)
                st.write()
                data={
                    "Nom":Nom,
                    "Date":date[0],
                    "type_p":type_pro,
                    "metrage":mots[3],
                    "kilometrage":mots[4]
                }
                collection.insert_one(data)


    

def alerte(message):
    st.markdown(f'<div style="background-color:#ffcccc;padding:10px;border-radius:5px;"><p style="color:red;">{message}</p></div>', unsafe_allow_html=True)
chemin_fichier = choisir_fichier()



import threading
import time

def replication_process():
    try:
        client_source = pymongo.MongoClient("mongodb://localhost:27017/")
        db_source = client_source["commandes"]
        collection_source = db_source["factures"]

        client_target = pymongo.MongoClient("mongodb://localhost:27018/")
        db_target = client_target["g_proconsulting"]
        collection_target = db_target["Facture_copies"]

        last_checked_timestamp = None

        while True:
            try:
                # Trouver les documents modifiés après le dernier timestamp vérifié
                query = {} if last_checked_timestamp is None else {"last_updated": {"$gt": last_checked_timestamp}}
                cursor = collection_source.find(query)

                # Répliquer les documents modifiés dans la collection cible
                for document in cursor:
                    collection_target.replace_one({"_id": document["_id"]}, document, upsert=True)

                # Récupérer les nouveaux documents ajoutés dans la collection source
                new_documents = collection_source.find({"_id": {"$gt": last_checked_timestamp}})

                # Insérer les nouveaux documents dans la collection cible
                for document in new_documents:
                    collection_target.insert_one(document)

                # Mettre à jour le timestamp de la dernière vérification
                last_checked_timestamp = time.time()

                # Attendre un certain temps avant de vérifier à nouveau
                time.sleep(5)  # Attendre 5 secondes avant de vérifier à nouveau
            except Exception as e:
                print(f"Une erreur s'est produite : {e}")
                # Gérer l'erreur, par exemple, en réessayant la connexion MongoDB
    except Exception as e:
        print(f"Une erreur s'est produite lors de l'exécution du thread : {e}")

# Lancer le processus de réplication dans un thread séparé

if type_facture :
    if chemin_fichier:
        informations_extraites = extraire_informations_document(chemin_fichier)
        st.write("Informations extraites :", informations_extraites)

        if informations_extraites:
            traitement(chemin_fichier,type_facture, informations_extraites)
        else:
             print("Aucune information extraite du fichier PDF.")
    else:
        print("Aucun fichier PDF sélectionné.")
#replication_thread = threading.Thread(target=replication_process)
#replication_thread.start()
